# views.py

from django.shortcuts import render, redirect, get_object_or_404
import openpyxl
from django.http import HttpResponse
from coupons.models import Coupon as CouponModel
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from accounts.models import OTPVerification, ProfileUser
from user.models import Address
from django.views.decorators.cache import never_cache
from django.utils import timezone
import random
import re
from django.core.paginator import Paginator
from django.db.models import Q
from accounts.email_utils import send_admin_otp_email
from django.utils.timezone import make_aware
from datetime import datetime, timedelta
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from order.models import Order, OrderItem
from product.models import Product, Variant
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from core.decorators import admin_required, unauthenticated_user


@never_cache
@unauthenticated_user
def admin_signin(request):

    # Already logged in admin
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_dashboard")
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        if not email:
            messages.error(request, "Email is required")
            return render(
                request, "auth/admin_signin.html", {"submitted_email": email}
            )
        if not password:
            messages.error(request, "password is required")
            return render(
                request, "auth/admin_signin.html", {"submitted_email": email}
            )

        user = authenticate(request, username=email, password=password)

        if user is None:
            messages.error(request, "Invalid email or password")
            return render(
                request, "auth/admin_signin.html", {"submitted_email": email}
            )

        # admin acces not normal usr
        if not user.is_superuser:
            messages.error(request, "Access denied.Admin only.")
            return render(
                request, "auth/admin_signin.html", {"submitted_email": email}
            )

        login(request, user)

        request.session["success"] = "Welcome to the Dashboard"
        return redirect("admin_dashboard")

    return render(request, "auth/admin_signin.html", {"submitted_email": ""})


@never_cache
@admin_required
def admin_dashboard(request):
    success = request.session.pop("success", None)

    filter_type = request.GET.get(
        "filter", "month"
    )  # Default to 'month' instead of 'all'
    now = timezone.localtime(timezone.now())
    today_midnight = now.replace(hour=0, minute=0, second=0)
    start_date = None
    end_date = now

    if filter_type == "day":
        start_date = today_midnight
    elif filter_type == "week":
        start_date = now - timedelta(days=7)
    elif filter_type == "month":
        start_date = now.replace(day=1, hour=0, minute=0, second=0)
    elif filter_type == "year":
        start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0)
    elif filter_type == "custom":
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        if start_str and end_str:
            try:
                start_date = make_aware(
                    datetime.strptime(start_str, "%Y-%m-%d")
                )
                end_date = make_aware(
                    datetime.strptime(end_str, "%Y-%m-%d")
                ).replace(hour=23, minute=59, second=59,)

                # Start must be before or equal to End
                if start_date > end_date:
                    messages.error(
                        request,
                        "Invalid Date Range: End date cannot be before start date.",  # noqa: E501
                    )
                    # Default back to month if validation fails
                    filter_type = "month"
                    start_date = now.replace(day=1, hour=0, minute=0, second=0)
                    end_date = now
            except (ValueError, TypeError):
                filter_type = "month"
                start_date = now.replace(day=1, hour=0, minute=0, second=0)
        else:
            filter_type = "month"
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now
    else:
        # If invalid filter, default to month
        filter_type = "month"
        start_date = now.replace(day=1, hour=0, minute=0, second=0)

    if filter_type == "custom":
        filtered_orders = Order.objects.filter(
            created_at__range=[start_date, end_date]
        )
    else:
        filtered_orders = Order.objects.filter(created_at__gte=start_date)

    total_revenue = (
        filtered_orders.aggregate(total=Sum("total_amount"))["total"] or 0
    )

    today_start = now.replace(hour=0, minute=0, second=0)
    today_revenue = (
        Order.objects.filter(created_at__gte=today_start).aggregate(
            total=Sum("total_amount")
        )["total"]
        or 0
    )

    month_start = now.replace(day=1, hour=0, minute=0, second=0)
    monthly_revenue = (
        Order.objects.filter(created_at__gte=month_start).aggregate(
            total=Sum("total_amount")
        )["total"]
        or 0
    )

    total_orders_count = filtered_orders.count()
    avg_order_value = (
        total_revenue / total_orders_count if total_orders_count > 0 else 0
    )

    pending_orders = filtered_orders.filter(order_status="PENDING").count()
    delivered_orders = filtered_orders.filter(order_status="DELIVERED").count()
    cancelled_orders = filtered_orders.filter(order_status="CANCELLED").count()
    return_orders = filtered_orders.filter(order_status="RETURNED").count()

    out_of_stock_variants = Variant.objects.filter(
        stock=0, is_active=True
    ).select_related("product")
    low_stock_variants = Variant.objects.filter(
        stock__lt=5, stock__gt=0, is_active=True
    ).select_related("product")
    out_of_stock_products = out_of_stock_variants.count()
    low_stock_products = low_stock_variants.count()

    total_users = ProfileUser.objects.filter(is_superuser=False).count()

    best_products_raw = (
        OrderItem.objects.filter(order__in=filtered_orders)
        .values("variant__product__id", "variant__product__name")
        .annotate(total_qty=Sum("quantity"))
        .order_by("-total_qty")[:10]
    )

    best_products = []
    for bp in best_products_raw:
        product_obj = Product.objects.get(id=bp["variant__product__id"])
        primary_image = product_obj.images.filter(is_primary=True).first()
        if not primary_image:
            primary_image = product_obj.images.first()
        image_url = primary_image.image.url if primary_image else None

        best_products.append(
            {
                "id": bp["variant__product__id"],
                "name": bp["variant__product__name"],
                "total_qty": bp["total_qty"],
                "image_url": image_url,
            }
        )

    best_categories = (
        OrderItem.objects.filter(order__in=filtered_orders)
        .values("variant__product__category__name")
        .annotate(total_qty=Sum("quantity"))
        .order_by("-total_qty")[:10]
    )

    recent_orders = filtered_orders.order_by("-created_at")[:10]

    # Generate chart data based on filter type
    chart_data = []

    if filter_type == "day":
        # Hourly data for today
        for hour in range(24):
            hour_start = now.replace(
                hour=hour, minute=0, second=0, microsecond=0
            )
            hour_end = hour_start + timedelta(hours=1)
            hour_revenue = (
                filtered_orders.filter(
                    created_at__gte=hour_start, created_at__lt=hour_end
                ).aggregate(total=Sum("total_amount"))["total"]
                or 0
            )

            chart_data.append(
                {"label": f"{hour:02d}:00", "total": float(hour_revenue)}
            )

    elif filter_type == "week":
        # Daily data for last 7 days
        for i in range(7):
            day = now - timedelta(days=6 - i)
            day_start = day.replace(hour=0, minute=0, second=0)
            day_end = day_start + timedelta(days=1)
            day_revenue = (
                filtered_orders.filter(
                    created_at__gte=day_start, created_at__lt=day_end
                ).aggregate(total=Sum("total_amount"))["total"]
                or 0
            )

            chart_data.append(
                {
                    "label": day.strftime("%a"),  # Mon, Tue, etc
                    "total": float(day_revenue),
                }
            )

    elif filter_type == "month":
        # Daily data for current month
        month_start = now.replace(day=1, hour=0, minute=0, second=0)
        days_in_month = (
            month_start.replace(month=month_start.month % 12 + 1, day=1)
            - timedelta(days=1)
        ).day

        for day in range(1, days_in_month + 1):
            day_date = month_start.replace(day=day)
            day_end = day_date + timedelta(days=1)
            day_revenue = (
                filtered_orders.filter(
                    created_at__gte=day_date, created_at__lt=day_end
                ).aggregate(total=Sum("total_amount"))["total"]
                or 0
            )

            chart_data.append({"label": str(day), "total": float(day_revenue)})

    elif filter_type == "year":
        # Monthly data for current year
        year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0)
        for month in range(1, 13):
            month_date = year_start.replace(month=month)
            if month == 12:
                month_end = year_start.replace(
                    year=year_start.year + 1, month=1, day=1
                )
            else:
                month_end = year_start.replace(month=month + 1, day=1)

            month_revenue = (
                filtered_orders.filter(
                    created_at__gte=month_date, created_at__lt=month_end
                ).aggregate(total=Sum("total_amount"))["total"]
                or 0
            )

            chart_data.append(
                {
                    "label": month_date.strftime("%b"),
                    "total": float(month_revenue),
                }
            )

    elif filter_type == "custom":
        delta = end_date - start_date
        if delta.days <= 3:
            # Hourly data for short ranges (up to 3 days) to make it "perfect"
            temp_date = start_date.replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            while temp_date <= end_date:
                for hour in [0, 4, 8, 12, 16, 20]:
                    h_start = temp_date + timedelta(hours=hour)
                    h_end = h_start + timedelta(hours=4)
                    if h_start > end_date:
                        continue
                    h_revenue = (
                        filtered_orders.filter(
                            created_at__gte=h_start, created_at__lt=h_end
                        ).aggregate(total=Sum("total_amount"))["total"]
                        or 0
                    )
                    chart_data.append(
                        {
                            "label": h_start.strftime("%b %d %H:%M"),
                            "total": float(h_revenue),
                        }
                    )
                temp_date += timedelta(days=1)
        elif delta.days <= 60:
            # Daily data for custom range (4 to 60 days)
            temp_date = start_date
            while temp_date <= end_date:
                day_start = temp_date.replace(
                    hour=0, minute=0, second=0, microsecond=0
                )
                day_end = day_start + timedelta(days=1)
                day_revenue = (
                    filtered_orders.filter(
                        created_at__gte=day_start, created_at__lt=day_end
                    ).aggregate(total=Sum("total_amount"))["total"]
                    or 0
                )
                chart_data.append(
                    {
                        "label": temp_date.strftime("%b %d"),
                        "total": float(day_revenue),
                    }
                )
                temp_date += timedelta(days=1)
        else:
            # Monthly data for ranges > 60 days
            temp_date = start_date.replace(
                day=1, hour=0, minute=0, second=0, microsecond=0
            )
            while temp_date <= end_date:
                m_start = temp_date
                if temp_date.month == 12:
                    m_end = temp_date.replace(year=temp_date.year + 1, month=1)
                else:
                    m_end = temp_date.replace(month=temp_date.month + 1)
                m_revenue = (
                    filtered_orders.filter(
                        created_at__gte=m_start, created_at__lt=m_end
                    ).aggregate(total=Sum("total_amount"))["total"]
                    or 0
                )
                chart_data.append(
                    {
                        "label": temp_date.strftime("%b %Y"),
                        "total": float(m_revenue),
                    }
                )
                temp_date = m_end
    else:
        # Default: show monthly data for current year
        chart_data_raw = (
            Order.objects.filter(created_at__year=now.year)
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(total=Sum("total_amount"))
            .order_by("month")
        )

        for entry in chart_data_raw:
            chart_data.append(
                {
                    "label": entry["month"].strftime("%b"),
                    "total": float(entry["total"]),
                }
            )

    context = {
        "success": success,
        "filter_type": filter_type,
        "total_revenue": total_revenue,
        "today_revenue": today_revenue,
        "monthly_revenue": monthly_revenue,
        "avg_order_value": avg_order_value,
        "total_orders": total_orders_count,
        "pending_orders": pending_orders,
        "delivered_orders": delivered_orders,
        "cancelled_orders": cancelled_orders,
        "return_orders": return_orders,
        "total_users": total_users,
        "out_of_stock_products": out_of_stock_products,
        "low_stock_products": low_stock_products,
        "out_of_stock_variants": out_of_stock_variants,
        "low_stock_variants": low_stock_variants,
        "best_products": best_products,
        "best_categories": best_categories,
        "recent_orders": recent_orders,
        "chart_data": chart_data,
        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "admin_dashboard.html", context)


@never_cache
def admin_signout(request):
    logout(request)
    return redirect("admin_signin")


@never_cache
@unauthenticated_user
def admin_forget_password(request):

    if request.method == "POST":
        email = request.POST.get("email")

        if not email:
            messages.error(request, "Email required")
            return render(
                request,
                "auth/admin_forget_password.html",
                {"submitted_email": email},
            )

        try:
            user = ProfileUser.objects.get(email=email)

            if not user.is_superuser:
                messages.error(request, "Not authorized")
                return render(
                    request,
                    "auth/admin_forget_password.html",
                    {"submitted_email": email},
                )
        except ProfileUser.DoesNotExist:
            messages.error(request, "Email not found")
            return render(
                request,
                "auth/admin_forget_password.html",
                {"submitted_email": email},
            )

        # if old otp have it will dlt
        OTPVerification.objects.filter(user=user).delete()
        otp = str(random.randint(1000, 9999))

        OTPVerification.objects.create(
            user=user,
            otp_code=otp,
            expires_at=timezone.now() + timedelta(minutes=2),
        )

        # Send premium admin OTP email
        send_admin_otp_email(user, otp)

        request.session["reset_user"] = user.id
        return redirect("admin_otp_verification")

    return render(
        request, "auth/admin_forget_password.html", {"submitted_email": ""}
    )


@never_cache
@unauthenticated_user
def admin_otp_verification(request):

    user_id = request.session.get("reset_user")

    if not user_id:
        return redirect("admin_forgot_password")

    user = ProfileUser.objects.get(id=user_id)

    if request.method == "POST":
        entered_otp = request.POST.get("otp")

        otp_obj = (
            OTPVerification.objects.filter(user=user)
            .order_by("-created_at")
            .first()
        )

        if not otp_obj:
            messages.error(request, "OTP not found")
            return redirect("admin_otp_verification")

        if timezone.now() > otp_obj.expires_at:
            messages.error(request, "OTP expired")
            return redirect("admin_forgot_password")

        if otp_obj.otp_code != entered_otp:
            messages.error(request, "Invalid OTP")
            return redirect("admin_otp_verification")

        otp_obj.is_verified = True
        otp_obj.save()

        otp_obj.delete()  # delte otp ot[ is one ytime use]

        request.session["otp_verified"] = (
            True  # delt otp aftr user cn still procced resset pass
        )
        return redirect("admin_reset_password")

    return render(request, "auth/admin_otp_verification.html")


@never_cache
@unauthenticated_user
def admin_resend_otp(request):

    user_id = request.session.get("reset_user")
    if not user_id:
        return redirect("admin_forgot_password")

    user = ProfileUser.objects.get(id=user_id)

    OTPVerification.objects.filter(user=user).delete()
    otp = str(random.randint(1000, 9999))

    OTPVerification.objects.create(
        user=user,
        otp_code=otp,
        expires_at=timezone.now() + timedelta(minutes=2),
    )

    # Send premium admin OTP email
    send_admin_otp_email(user, otp)
    messages.success(request, "New OTP sent")
    return redirect("admin_otp_verification")


@never_cache
@unauthenticated_user
def admin_reset_password(request):

    user_id = request.session.get("reset_user")
    verified = request.session.get("otp_verified")

    if not user_id or not verified:
        return redirect("admin_forgot_password")

    user = ProfileUser.objects.get(id=user_id)

    if request.method == "POST":
        password = request.POST.get("password")
        confirm = request.POST.get("confirm_password")

        if not password or not confirm:
            messages.error(request, "All fields required")
            return redirect("admin_reset_password")

        if password != confirm:
            messages.error(request, "Passwords do not match")
            return redirect("admin_reset_password")

        pattern_pass = r"^(?=.*[A-Z])(?=.*[0-9])(?=.*[!@#$%^&*]).{8,}$"
        if not re.match(pattern_pass, password):
            messages.error(request, "Weak password")
            return redirect("admin_reset_password")

        user.set_password(password)
        user.save()

        # cleanup
        OTPVerification.objects.filter(user=user).delete()
        request.session.flush()

        messages.success(request, "Password reset successful")
        return redirect("admin_signin")

    return render(request, "auth/admin_reset_password.html")


@never_cache
@admin_required
def user_management(request):
    q = request.GET.get("q", "")
    status = request.GET.get("status", "")

    users = ProfileUser.objects.filter(is_superuser=False).order_by(
        "-date_joined"
    )

    if q:
        users = users.filter(Q(full_name__icontains=q) | Q(email__icontains=q))
    if status == "active":
        users = users.filter(is_active=True)
    elif status == "blocked":
        users = users.filter(is_active=False)

    paginator = Paginator(users, 5)
    page = request.GET.get("page", 1)
    page_obj = paginator.get_page(page)

    context = {
        "users": page_obj,
        "page_obj": page_obj,
        "q": q,
        "status": status,
    }

    return render(request, "user_management.html", context)


@never_cache
@admin_required
def admin_toggle_block(request, id):
    if request.method != "POST":
        return redirect("user_management")
    user = get_object_or_404(ProfileUser, id=id)

    # Toggle is_activ

    user.is_active = not user.is_active
    user.save(update_fields=["is_active"])

    name = user.get_full_name() or user.email
    action = "unblocked" if user.is_active else "blocked"
    messages.success(request, f"{name} has been {action} successfully.")

    return redirect("user_management")


@never_cache
@admin_required
def user_detail(request, id):
    # Get the user
    user = get_object_or_404(ProfileUser, id=id, is_superuser=False)

    # Get all orders for this user
    all_orders = Order.objects.filter(user=user).order_by("-created_at")

    # Calculate total orders
    total_orders = all_orders.count()

    total_spent = 0
    for order in all_orders:
        total_spent += float(order.total_amount)

    # Calculate average order value
    if total_orders > 0:
        avg_order = round(total_spent / total_orders, 2)
    else:
        avg_order = 0

    # Count delivered orders
    delivered_orders = all_orders.filter(order_status="DELIVERED").count()

    # Paginate
    paginator = Paginator(all_orders, 5)
    page_number = request.GET.get("page", 1)
    orders_page = paginator.get_page(page_number)

    # Get user's addresses
    all_addresses = Address.objects.filter(user=user)

    # Paginate addresses
    address_paginator = Paginator(all_addresses, 3)
    address_page_number = request.GET.get("address_page", 1)
    addresses_page = address_paginator.get_page(address_page_number)

    phone_number = user.phone_number if user.phone_number else "Not provided"
    full_name = user.get_full_name() if user.get_full_name() else user.username
    joined_date = user.date_joined

    # Get user's last login
    last_login = user.last_login if user.last_login else "Never"

    context = {
        "user": user,
        "total_orders": total_orders,
        "total_spent": total_spent,
        "avg_order": avg_order,
        "delivered_count": delivered_orders,
        "orders": orders_page,
        "addresses": addresses_page,
        "phone_number": phone_number,
        "full_name": full_name,
        "joined_date": joined_date,
        "last_login": last_login,
    }

    return render(request, "user_detail.html", context)


@never_cache
@admin_required
def sales_report(request):

    filter_type = request.GET.get("filter", "month")
    now = timezone.localtime(timezone.now())
    today_midnight = now.replace(hour=0, minute=0, second=0)

    if filter_type == "day":
        # Today (since midnight)
        start_date = today_midnight
        end_date = now
    elif filter_type == "week":
        # Last 7 days
        start_date = now - timedelta(days=7)
        end_date = now
    elif filter_type == "year":
        # Last 365 days
        start_date = now - timedelta(days=365)
        end_date = now
    elif filter_type == "custom":
        start = request.GET.get("start_date")
        end = request.GET.get("end_date")
        if start and end:
            try:
                start_date = make_aware(datetime.strptime(start, "%Y-%m-%d"))
                end_date = make_aware(
                    datetime.strptime(end, "%Y-%m-%d")
                ).replace(hour=23, minute=59, second=59)
                
                #End date cannot be before start date
                if end_date < start_date:
                    messages.error(
                        request,
                        "Invalid Date Range: End date cannot be before start date.",
                    )
                    # Default back to month if validation fails
                    filter_type = "month"
                    start_date = now.replace(day=1, hour=0, minute=0, second=0)
                    end_date = now
                #Start date cannot be in the future
                elif start_date > now:
                    messages.error(
                        request,
                        "Invalid Date Range: Start date cannot be in the future.",
                    )
                    filter_type = "month"
                    start_date = now.replace(day=1, hour=0, minute=0, second=0)
                    end_date = now
                #End date cannot be in the future
                elif end_date > now:
                    messages.error(
                        request,
                        "Invalid Date Range: End date cannot be in the future.",
                    )
                    filter_type = "month"
                    start_date = now.replace(day=1, hour=0, minute=0, second=0)
                    end_date = now
                #Date range cannot exceed 1 year
                elif (end_date - start_date).days > 365:
                    messages.error(
                        request,
                        "Invalid Date Range: Date range cannot exceed 1 year.",
                    )
                    filter_type = "month"
                    start_date = now.replace(day=1, hour=0, minute=0, second=0)
                    end_date = now
            except (ValueError, TypeError):
                messages.error(
                    request,
                    "Invalid Date Format: Please use valid dates.",
                )
                filter_type = "month"
                start_date = now.replace(day=1, hour=0, minute=0, second=0)
                end_date = now
        else:
            messages.error(
                request,
                "Missing Dates: Both start and end dates are required.",
            )
            filter_type = "month"
            start_date = now.replace(day=1, hour=0, minute=0, second=0)
            end_date = now
    else:  # default: month
        # Start from the first day of the current month
        start_date = now.replace(day=1, hour=0, minute=0, second=0)
        end_date = now

    # Get all delivered orders in the date range
    orders = Order.objects.filter(
        created_at__range=[start_date, end_date], order_status="DELIVERED"
    )

    total_revenue = 0.0
    coupon_discount = 0.0
    for order in orders:
        total_revenue += float(order.total_amount)
        coupon_discount += float(order.discount_amount)

    # Count total products sold
    order_items = OrderItem.objects.filter(order__in=orders)
    products_sold = sum(item.quantity for item in order_items)

    chart_dict = {}
    chart_list = []
    
    if filter_type == "day":
        # Hourly data for Today - show all 24 hours
        for h in range(24):
            hour_key = f"{h:02d}:00"
            chart_dict[hour_key] = 0.0

        # Aggregate orders by hour
        for order in orders:
            hour_key = order.created_at.strftime("%H:00")
            if hour_key in chart_dict:
                chart_dict[hour_key] += float(order.total_amount)
        
        # Create ordered list
        chart_list = [{"date": f"{h:02d}:00", "total": chart_dict[f"{h:02d}:00"]} for h in range(24)]
        
    elif filter_type == "custom":
        # For custom dates, determine granularity based on range
        delta_days = (end_date - start_date).days
        
        if delta_days <= 1:
            # Hourly data for 1 day or less - show all 24 hours
            for h in range(24):
                hour_key = f"{h:02d}:00"
                chart_dict[hour_key] = 0.0
            
            for order in orders:
                hour_key = order.created_at.strftime("%H:00")
                if hour_key in chart_dict:
                    chart_dict[hour_key] += float(order.total_amount)
            
            # Create ordered list
            chart_list = [{"date": f"{h:02d}:00", "total": chart_dict[f"{h:02d}:00"]} for h in range(24)]
            
        elif delta_days <= 60:
            # Daily data for ranges up to 60 days
            temp_date = start_date.date()
            end_date_only = end_date.date()
            date_list = []
            
            while temp_date <= end_date_only:
                formatted_date = temp_date.strftime("%b %d")
                chart_dict[str(temp_date)] = 0.0
                date_list.append((str(temp_date), formatted_date))
                temp_date += timedelta(days=1)
            
            for order in orders:
                date_key = str(order.created_at.date())
                if date_key in chart_dict:
                    chart_dict[date_key] += float(order.total_amount)
            
            # Create ordered list with formatted dates
            chart_list = [{"date": formatted, "total": chart_dict[date_key]} for date_key, formatted in date_list]
            
        else:
            # Monthly data for ranges > 60 days
            temp_date = start_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_list = []
            
            while temp_date <= end_date:
                month_key = temp_date.strftime("%Y-%m")
                month_label = temp_date.strftime("%b %Y")
                chart_dict[month_key] = 0.0
                month_list.append((month_key, month_label))
                
                # Move to next month
                if temp_date.month == 12:
                    temp_date = temp_date.replace(year=temp_date.year + 1, month=1)
                else:
                    temp_date = temp_date.replace(month=temp_date.month + 1)
            
            for order in orders:
                month_key = order.created_at.strftime("%Y-%m")
                if month_key in chart_dict:
                    chart_dict[month_key] += float(order.total_amount)
            
            # Create ordered list with formatted labels
            chart_list = [{"date": label, "total": chart_dict[key]} for key, label in month_list]
            
    else:
        # Daily data for week, month, year filters
        temp_date = start_date.date()
        end_date_only = end_date.date()
        
        if filter_type == "week":
            # Show day name for week view (Mon, Tue, etc.)
            date_list = []
            while temp_date <= end_date_only:
                day_name = temp_date.strftime("%a")
                chart_dict[str(temp_date)] = 0.0
                date_list.append((str(temp_date), day_name))
                temp_date += timedelta(days=1)
            
            for order in orders:
                date_key = str(order.created_at.date())
                if date_key in chart_dict:
                    chart_dict[date_key] += float(order.total_amount)
            
            # Create ordered list
            chart_list = [{"date": day_name, "total": chart_dict[date_key]} for date_key, day_name in date_list]
            
        elif filter_type == "month":
            # Show day number for month view (1, 2, 3, etc.)
            date_list = []
            while temp_date <= end_date_only:
                day_num = str(temp_date.day)
                chart_dict[str(temp_date)] = 0.0
                date_list.append((str(temp_date), day_num))
                temp_date += timedelta(days=1)
            
            for order in orders:
                date_key = str(order.created_at.date())
                if date_key in chart_dict:
                    chart_dict[date_key] += float(order.total_amount)
            
            # Create ordered list
            chart_list = [{"date": day_num, "total": chart_dict[date_key]} for date_key, day_num in date_list]
            
        elif filter_type == "year":
            # Show last 12 months from today (rolling 12 months)
            month_list = []
            
            # Start from 12 months ago
            temp_date = (now - timedelta(days=365)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # Generate 12 months
            for i in range(12):
                month_key = temp_date.strftime("%Y-%m")
                month_name = temp_date.strftime("%b %y")  # e.g., "May 25", "Jun 25"
                chart_dict[month_key] = 0.0
                month_list.append((month_key, month_name))
                
                # Move to next month
                if temp_date.month == 12:
                    temp_date = temp_date.replace(year=temp_date.year + 1, month=1)
                else:
                    temp_date = temp_date.replace(month=temp_date.month + 1)
            
            for order in orders:
                month_key = order.created_at.strftime("%Y-%m")
                if month_key in chart_dict:
                    chart_dict[month_key] += float(order.total_amount)
            
            # Create ordered list
            chart_list = [{"date": month_name, "total": chart_dict[month_key]} for month_key, month_name in month_list]
            
        else:
            # Default: show full date
            date_list = []
            while temp_date <= end_date_only:
                date_str = str(temp_date)
                chart_dict[date_str] = 0.0
                date_list.append(date_str)
                temp_date += timedelta(days=1)
            
            for order in orders:
                date_str = str(order.created_at.date())
                if date_str in chart_dict:
                    chart_dict[date_str] += float(order.total_amount)
            
            # Create ordered list
            chart_list = [{"date": date_str, "total": chart_dict[date_str]} for date_str in date_list]

    # Calculate previous period for growth comparison
    prev_start = None
    prev_end = None
    
    if filter_type == "day":
        # Exactly yesterday (full day)
        prev_date = start_date.date() - timedelta(days=1)
        prev_start = timezone.make_aware(
            datetime.combine(prev_date, datetime.min.time())
        )
        prev_end = timezone.make_aware(
            datetime.combine(prev_date, datetime.max.time())
        )
        prev_rev_sum = (
            Order.objects.filter(
                created_at__date=prev_date, order_status="DELIVERED"
            ).aggregate(total=Sum("total_amount"))["total"]
            or 0
        )
        previous_revenue = float(prev_rev_sum)
    else:

        if filter_type == "week":
            prev_start, prev_end = start_date - timedelta(days=7), start_date
        elif filter_type == "year":
            prev_start, prev_end = start_date - timedelta(days=365), start_date
        elif filter_type == "month":
            comparison_type = request.GET.get("comparison", "partial")
            last_month_end = start_date - timedelta(seconds=1)
            prev_start = last_month_end.replace(
                day=1, hour=0, minute=0, second=0
            )
            
            if comparison_type == "partial":
                # Compare current month-to-date with same days in previous month
                # e.g., May 1-2 vs April 1-2
                try:
                    prev_end = prev_start.replace(
                        day=now.day,
                        hour=now.hour,
                        minute=now.minute,
                        second=now.second,
                    )
                except ValueError:
                    # Handle cases like March 31 -> Feb 28
                    prev_end = last_month_end
            else:
                # Compare current month-to-date with FULL previous month
                prev_end = last_month_end
        elif filter_type == "custom":
            # For custom dates, calculate the previous period with same duration
            period_length = end_date - start_date
            # Previous period ends where current period starts (minus 1 second)
            prev_end = start_date - timedelta(seconds=1)
            # Previous period starts by going back the same duration
            prev_start = prev_end - period_length
            # Ensure prev_start has proper time set (start of day)
            prev_start = prev_start.replace(hour=0, minute=0, second=0)
        else:
            # Default fallback for any other filter type
            period_length = end_date - start_date
            prev_start, prev_end = start_date - period_length, start_date

        prev_rev_sum = (
            Order.objects.filter(
                created_at__range=[prev_start, prev_end],
                order_status="DELIVERED",
            ).aggregate(total=Sum("total_amount"))["total"]
            or 0
        )
        previous_revenue = float(prev_rev_sum)

    # Growth percentage calculation
    if previous_revenue > 0:
        growth = round(
            ((total_revenue - previous_revenue) / previous_revenue) * 100, 2
        )
    else:
        growth = 100.0 if total_revenue > 0 else 0.0

    orders_list = orders.select_related("payment").order_by("-created_at")
    coupon_stats = CouponModel.objects.filter(
        used_count__gt=0, is_deleted=False
    ).order_by("-used_count")

    order_page_obj = Paginator(orders_list, 4).get_page(
        request.GET.get("order_page")
    )
    coupon_page_obj = Paginator(coupon_stats, 4).get_page(
        request.GET.get("coupon_page")
    )

    context = {
        "total_revenue": total_revenue,
        "previous_revenue": previous_revenue,
        "total_orders": orders.count(),
        "products_sold": products_sold,
        "coupon_discount": coupon_discount,
        "chart_data": chart_list,
        "growth": growth,
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
        "prev_start_date": prev_start,
        "prev_end_date": prev_end,
        "comparison_type": request.GET.get("comparison", "partial"),
        "orders_list": order_page_obj,
        "coupon_stats": coupon_page_obj,
    }

    return render(request, "admin/sales_report.html", context)


@never_cache
@admin_required
def export_sales_excel(request):
    filter_type = request.GET.get("filter", "month")
    now = timezone.localtime(timezone.now())
    today_midnight = now.replace(hour=0, minute=0, second=0)

    if filter_type == "day":
        start_date, end_date = today_midnight, now
    elif filter_type == "week":
        start_date, end_date = now - timedelta(days=7), now
    elif filter_type == "year":
        start_date, end_date = now - timedelta(days=365), now
    elif filter_type == "custom":
        start, end = request.GET.get("start_date"), request.GET.get("end_date")
        if start and end:
            try:
                start_date = make_aware(datetime.strptime(start, "%Y-%m-%d"))
                end_date = make_aware(datetime.strptime(end, "%Y-%m-%d")).replace(
                    hour=23, minute=59, second=59
                )
                # Validation
                if end_date < start_date or start_date > now or end_date > now:
                    start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0,), now
            except Exception:
                start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now
        else:
            start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now
    else:
        start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now

    orders = (
        Order.objects.filter(
            created_at__range=[start_date, end_date], order_status="DELIVERED"
        )
        .select_related("payment")
        .order_by("-created_at")
    )

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_fill = PatternFill(
        start_color="4A9050", end_color="4A9050", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center")
    money_format = "₹#,##0.00"
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Order Number",
        "Date",
        "Subtotal",
        "Discount",
        "Total Amount",
        "Payment Method",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    total_sum = 0
    for order in orders:
        pm = getattr(order.payment, "payment_method", "N/A")
        ws.append(
            [
                order.order_number,
                order.created_at.strftime("%Y-%m-%d"),
                float(order.subtotal),
                float(order.discount_amount),
                float(order.total_amount),
                pm,
            ]
        )
        total_sum += float(order.total_amount)
        curr_row = ws.max_row
        for i, cell in enumerate(ws[curr_row], 1):
            cell.border = thin_border
            if i in [3, 4, 5]:
                cell.number_format = money_format
            if i in [1, 2, 6]:
                cell.alignment = center_align

    ws.append([])
    ws.append(["", "", "", "GRAND TOTAL", total_sum])
    total_row = ws.max_row
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = money_format

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # noqa: E501
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Glowe_Sales_{filter_type}.xlsx"'
    )
    wb.save(response)
    return response



@never_cache
@admin_required
def export_sales_pdf(request):
    """
    Generates a modern, professional Sales Report PDF.
    Includes: Order Number, Date, Qty, Subtotal, Discount, Payment Method, Total.
    """
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import HRFlowable, KeepTogether

    # ── 1. Resolve date range ──────────────────────────────────────────────
    filter_type = request.GET.get("filter", "month")
    now = timezone.localtime(timezone.now())
    today_midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)

    if filter_type == "day":
        start_date, end_date = today_midnight, now
    elif filter_type == "week":
        start_date, end_date = now - timedelta(days=7), now
    elif filter_type == "year":
        start_date, end_date = now - timedelta(days=365), now
    elif filter_type == "custom":
        start, end = request.GET.get("start_date"), request.GET.get("end_date")
        if start and end:
            try:
                start_date = make_aware(datetime.strptime(start, "%Y-%m-%d"))
                end_date = make_aware(datetime.strptime(end, "%Y-%m-%d")).replace(
                    hour=23, minute=59, second=59
                )
                if end_date < start_date or start_date > now or end_date > now:
                    start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now
            except Exception:
                start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now
        else:
            start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now
    else:  # month (default)
        start_date, end_date = now.replace(day=1, hour=0, minute=0, second=0), now

    # ── 2. Fetch data ──────────────────────────────────────────────────────
    orders = (
        Order.objects.filter(
            created_at__range=[start_date, end_date],
            order_status="DELIVERED",
        )
        .select_related("payment")
        .prefetch_related("items")
        .order_by("-created_at")
    )

    total_orders    = orders.count()
    total_revenue   = sum(float(o.total_amount)    for o in orders)
    total_discount  = sum(float(o.discount_amount) for o in orders)
    total_qty       = sum(
        sum(item.quantity for item in o.items.all()) for o in orders
    )

    # ── 3. HTTP response ──────────────────────────────────────────────────
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="Glowe_Sales_{filter_type}.pdf"'
    )

    # ── 4. Document setup ──────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    # ── 5. Colour palette ──────────────────────────────────────────────────
    C_DARK      = colors.HexColor("#0f1923")   # deep navy-black
    C_GREEN     = colors.HexColor("#16a34a")   # brand green
    C_LIGHT_BG  = colors.HexColor("#f0fdf4")   # very light green tint
    C_STRIPE    = colors.HexColor("#f8fafc")   # alternate row
    C_BORDER    = colors.HexColor("#e2e8f0")   # subtle border
    C_MUTED     = colors.HexColor("#64748b")   # muted text
    C_WHITE     = colors.white
    C_HEADER_BG = colors.HexColor("#1e3a2f")   # dark green header

    # ── 6. Paragraph styles ────────────────────────────────────────────────
    brand_style = ParagraphStyle(
        "brand", fontName="Helvetica-Bold", fontSize=26,
        textColor=C_DARK, leading=30,
    )
    report_title_style = ParagraphStyle(
        "report_title", fontName="Helvetica-Bold", fontSize=11,
        textColor=C_MUTED, leading=14, spaceAfter=2,
    )
    period_style = ParagraphStyle(
        "period", fontName="Helvetica", fontSize=9,
        textColor=C_MUTED, leading=12,
    )
    footer_style = ParagraphStyle(
        "footer", fontName="Helvetica", fontSize=8,
        textColor=C_MUTED, alignment=TA_CENTER,
    )
    summary_label_style = ParagraphStyle(
        "slabel", fontName="Helvetica", fontSize=8,
        textColor=C_MUTED, alignment=TA_CENTER, leading=10,
    )
    summary_value_style = ParagraphStyle(
        "svalue", fontName="Helvetica-Bold", fontSize=14,
        textColor=C_DARK, alignment=TA_CENTER, leading=18,
    )

    # ── 7. Build flowables ─────────────────────────────────────────────────
    elements = []

    # --- HEADER ACCENT BAR ---
    elements.append(HRFlowable(
        width="100%", thickness=5, color=C_GREEN, spaceAfter=14,
    ))

    # --- BRAND + PERIOD HEADER TABLE ---
    period_str = (
        f"{start_date.strftime('%d %b %Y')}  →  {end_date.strftime('%d %b %Y')}"
    )
    filter_label = {
        "day": "Daily Report", "week": "Weekly Report",
        "month": "Monthly Report", "year": "Annual Report",
        "custom": "Custom Period Report",
    }.get(filter_type, "Sales Report")

    header_data = [[
        Paragraph("GLOWÉ", brand_style),
        Table(
            [
                [Paragraph(filter_label.upper(), report_title_style)],
                [Paragraph(f"Period: {period_str}", period_style)],
                [Paragraph(f"Generated: {now.strftime('%d %b %Y, %I:%M %p')}", period_style)],
            ],
            colWidths=[280],
        ),
    ]]
    header_table = Table(header_data, colWidths=[190, 320])
    header_table.setStyle(TableStyle([
        ("VALIGN",  (0, 0), (-1, -1), "TOP"),
        ("ALIGN",   (1, 0), (1, 0),   "RIGHT"),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 18))

    # --- SUMMARY CARDS (4 metrics) ---
    def summary_cell(label, value):
        return Table(
            [
                [Paragraph(label, summary_label_style)],
                [Paragraph(value, summary_value_style)],
            ],
            colWidths=[116],
            rowHeights=[14, 28],
        )

    card_table = Table(
        [[
            summary_cell("TOTAL ORDERS",   str(total_orders)),
            summary_cell("TOTAL REVENUE",  f"Rs.{total_revenue:,.2f}"),
            summary_cell("TOTAL DISCOUNT", f"Rs.{total_discount:,.2f}"),
            summary_cell("UNITS SOLD",     str(total_qty)),
        ]],
        colWidths=[118, 118, 118, 118],
    )
    card_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_LIGHT_BG),
        ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [4]),
    ]))
    elements.append(card_table)
    elements.append(Spacer(1, 22))

    # --- SECTION TITLE ---
    elements.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=6))
    elements.append(Paragraph("Order Details", ParagraphStyle(
        "sec_title", fontName="Helvetica-Bold", fontSize=10,
        textColor=C_DARK, spaceBefore=2, spaceAfter=8,
    )))

    # --- MAIN ORDERS TABLE ---
    col_headers = [
        "#", "Order Number", "Date", "Qty",
        "Subtotal", "Discount", "Payment", "Total"
    ]
    col_widths = [20, 110, 68, 28, 68, 58, 60, 68]

    table_data = [col_headers]

    for idx, o in enumerate(orders, start=1):
        qty = sum(item.quantity for item in o.items.all())
        pm  = getattr(o.payment, "payment_method", "N/A")
        # Make payment method friendly
        pm_display = {
            "COD":    "Cash",
            "ONLINE": "Online",
            "WALLET": "Wallet",
        }.get(pm, pm)

        table_data.append([
            str(idx),
            o.order_number,
            o.created_at.strftime("%d %b %Y"),
            str(qty),
            f"Rs.{float(o.subtotal):,.2f}",
            f"Rs.{float(o.discount_amount):,.2f}",
            pm_display,
            f"Rs.{float(o.total_amount):,.2f}",
        ])

    # Grand Total row
    table_data.append([
        "", "GRAND TOTAL", "", str(total_qty),
        "", f"Rs.{total_discount:,.2f}", "",
        f"Rs.{total_revenue:,.2f}",
    ])

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    last_row = len(table_data) - 1

    main_table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),        C_HEADER_BG),
        ("TEXTCOLOR",     (0, 0), (-1, 0),        C_WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),        "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),        8),
        ("ALIGN",         (0, 0), (-1, 0),        "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, 0),        8),
        ("BOTTOMPADDING", (0, 0), (-1, 0),        8),

        # Data rows
        ("FONTNAME",      (0, 1), (-1, last_row - 1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, last_row - 1), 8),
        ("TOPPADDING",    (0, 1), (-1, last_row - 1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, last_row - 1), 5),

        # Zebra stripes
        ("ROWBACKGROUNDS", (0, 1), (-1, last_row - 1), [C_WHITE, C_STRIPE]),

        # Number column — centre
        ("ALIGN",         (0, 1), (0, last_row - 1), "CENTER"),
        # Qty column — centre
        ("ALIGN",         (3, 1), (3, last_row - 1), "CENTER"),
        # Payment — centre
        ("ALIGN",         (6, 1), (6, last_row - 1), "CENTER"),
        # Money columns — right
        ("ALIGN",         (4, 1), (4, last_row - 1), "RIGHT"),
        ("ALIGN",         (5, 1), (5, last_row - 1), "RIGHT"),
        ("ALIGN",         (7, 1), (7, last_row - 1), "RIGHT"),

        # Grid
        ("GRID",          (0, 0), (-1, last_row - 1), 0.4, C_BORDER),
        ("LINEBELOW",     (0, last_row - 1), (-1, last_row - 1), 0.8, C_DARK),

        # Grand Total row
        ("BACKGROUND",    (0, last_row), (-1, last_row), C_LIGHT_BG),
        ("FONTNAME",      (0, last_row), (-1, last_row), "Helvetica-Bold"),
        ("FONTSIZE",      (0, last_row), (-1, last_row), 9),
        ("TEXTCOLOR",     (0, last_row), (-1, last_row), C_DARK),
        ("ALIGN",         (1, last_row), (1, last_row),  "LEFT"),
        ("ALIGN",         (3, last_row), (3, last_row),  "CENTER"),
        ("ALIGN",         (5, last_row), (5, last_row),  "RIGHT"),
        ("ALIGN",         (7, last_row), (7, last_row),  "RIGHT"),
        ("TOPPADDING",    (0, last_row), (-1, last_row), 8),
        ("BOTTOMPADDING", (0, last_row), (-1, last_row), 8),
        ("BOX",           (0, last_row), (-1, last_row), 0.8, C_GREEN),
    ]))

    elements.append(main_table)
    elements.append(Spacer(1, 24))

    # --- FOOTER ---
    elements.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=6))
    elements.append(Paragraph(
        f"GLOWÉ · Sales Report · {now.strftime('%d %b %Y, %I:%M %p')} · Confidential",
        footer_style,
    ))

    # ── 8. Build PDF ──────────────────────────────────────────────────────
    doc.build(elements)
    return response
